#!/usr/bin/env python
# coding: utf-8

# ### 에듀넷 자동 회원가입
# - 에듀넷 사이트 접속
# - 회원가입 하기
# - 아이디 : 성명 이니셜 + 생년월일 8개
# - 학부모 연락처, 이메일 필요

# In[75]:


from selenium import webdriver
import time
import requests
from selenium.webdriver.common.keys import Keys
import openpyxl
import pandas as pd


# In[76]:


url = 'https://st.edunet.net/member/memberJoin?in_div=nedu'
driver = webdriver.Chrome()
v_time = 1
v_year = 2018
path = './data/edunet_21.xlsx'


# ## 1단계 - 엑셀파일에서 데이터 추출하기

# #### 사전 만들기 - 학생, 학부모, 아이디 정보

# In[79]:


def load_excel_file(p_path):
    #엑셀파일 불러오기
    wb = openpyxl.load_workbook(p_path)

    #시트 이름 확인하기
    wb.sheetnames

    # 데이터가 든 시트 선택하기 
    sheet = wb['Sheet1']

    # 마지막행 확인하기
    last_row = sheet.max_row
    
    return sheet, last_row


# In[84]:


# 함수 - 학생정보 딕셔너리
def make_dic_student(p_list):
    print('학생 정보 생성 중...')
    for i in range(2, last_row+1):
        birth = sheet.cell(row = i, column = 4).value
        dic = {
            'name' : sheet.cell(row = i, column = 2).value,
            'gen' : sheet.cell(row = i, column = 3).value,
            'birthY' : str(birth)[0: 4],
            'birthM' : str(birth)[4: 6],
            'birthD' : str(birth)[6:]
        }
        p_list.append(dic)
    return p_list


# In[85]:


# 함수 - 학부모정보 딕셔너리
def make_dic_parent(p_list):
    for i in range(2, last_row+1):
        tel = sheet.cell(row = i, column = 7).value
        dic = {
            'proname2' : sheet.cell(row = i, column = 5).value,
            'gen' : sheet.cell(row = i, column = 6).value,
            'parentTel1' : str(tel)[0:3],
            'parentTel2' : str(tel)[4:8],
            'parentTel3' : str(tel)[7:]
        }
        p_list.append(dic)
    return p_list


# In[86]:


# 함수 - 회원가입 정보 딕셔너리 
def make_dic_creat_Id(p_list):
    print('회원가입 정보 생성 중...')
    for i in range(2, last_row+1):
        email = sheet.cell(row = i, column = 9).value
        dic = {
            'userId' : sheet.cell(row = i, column = 8).value,
            'password' : 'h!12345678', 
            'repwd' : 'h!12345678',
            'subEmail1' : email.split('@')[0],
            'subEmail2' : email.split('@')[1]
        }
        p_list.append(dic)
    return p_list


# In[87]:


# 실행문

#엑셀부르기
sheet, last_row = load_excel_file(path)

#리스트 선언
list_student = []
list_parent = []
list_creat_id = []

#딕셔너리 만들기
print('홈페이지 가입 정보 생성중..')
student = make_dic_student(list_student)
parent = make_dic_parent(list_parent)
creat_id = make_dic_creat_Id(list_creat_id)


# ## 2단계 - 데이터를 이용하여 회원가입 하기(자동화)

# #### 에듀넷 사이트 접속

# In[88]:


def move_page(p_url):    
    print('웹페이지로 이동중...')
    driver.get(url)
    time.sleep(v_time)


# #### 회원가입하기 - 약관 동의

# In[89]:


list_agree = ['//*[@id="agr1"]', '//*[@id="agr2"]', '//*[@id="mem_sub_list"]/div[1]/fieldset/div[3]/div[2]/a[1]']

def agree_asignment(p_list_agree):
    for i in p_list_agree:
        driver.find_element_by_xpath("%s" %(i)).click()  #이용약관 동의 체크박스 클릭
        time.sleep(v_time)


# #### 회원가입하기 - 인증 및 회원정보 입력

# In[90]:


def fillout_xpath(p_dic_any):
    for k in p_dic_any:
        driver.find_element_by_xpath('//*[@id="%s"]' %(k)).send_keys('%s' %(p_dic_any[k])) 
        time.sleep(v_time)


# #### 함수 - 학생정보 채우기

# In[91]:


def fillout_student_info(p_dic_person_info):
    # print('%s 회원가입 시작!' %(p_dic_person_info['name']))
    #성별을 판단하여 클릭을 미리하고 요소 삭제
    if p_dic_person_info['gen'] == '남':
        gen = 'male'
    else:
        gen = 'female'
    driver.find_element_by_xpath('//*[@id="%s"]' %(gen)).click()
    del p_dic_person_info['gen'] #딕셔너리에서 성별 정보 삭제
    time.sleep(v_time)
    
    fillout_xpath(p_dic_person_info)


# #### 함수 - 학부모정보 채우기

# In[92]:


def fillout_parent_info(p_dic_person_info): #학부모 정보 채우기
    #성별을 판단하여 클릭을 미리하고 요소 삭제
    if p_dic_person_info['gen'] == '남':
        gen = 0
    else:
        gen = 1
    driver.find_elements_by_css_selector("input#progender2")[gen].click()
    del p_dic_person_info['gen'] #딕셔너리에서 성별 정보 삭제
    time.sleep(v_time)
    
    fillout_xpath(p_dic_person_info)


# In[93]:


def popup_accept():
    alert = driver.switch_to_alert() #팝업창을 변수로 지정
    alert.accept() #팝업창 확인 클릭


# In[94]:


def fillout_all_info(p_dic_student, p_dic_parent):
    fillout_student_info(p_dic_student)
    driver.find_element_by_xpath('//*[@id="liByAccr"]/a').click() #오프라인 동의 클릭

    popup_accept() #팝업창 수락 클릭후 닫기

    fillout_parent_info(p_dic_parent)
    driver.find_element_by_xpath('//*[@id="agrAll2"]').click() # 개인정보 처리 및 오프라인 가입 동의 클릭


# #### 실행문

# In[95]:


for i in range(last_row-1):
    # 회원가입 페이지 이동
    move_page(url)
    driver.implicitly_wait(5)

    #약관동의
    agree_asignment(list_agree)
    driver.implicitly_wait(5)
    print('홈페이지 약관동의 완료!')

    # 학생 학부모 개인정보 입력
    print('%s 홈페이지 가입 시작!' %(student[i]['name']) )
    fillout_all_info(student[i], parent[i])

    # 아이디 및 비밀번호 생성
    print('아이디와 비밀번호 생성 시작..')
    fillout_xpath(creat_id[i]) #아이디, 비번 채우기
    driver.find_element_by_xpath('//*[@id="dpChk"]').click() #중복확인

    popup_accept() #중복확인창 닫기

    # 학교기관 및 학년 입력
    print('소속기관 및 학년 입력 시작..')
    driver.find_element_by_xpath('//*[@id="btnSearch"]').click() #학교선택 버튼 클릭
    time.sleep(v_time)
    driver.find_element_by_xpath('//*[@id="schl_name"]').send_keys('초등학교') #학교급 선택
    time.sleep(v_time)
    driver.find_element_by_xpath('//*[@id="popupSchInput"]').send_keys('학현초등학교') #학교급 선택
    time.sleep(v_time)
    driver.find_element_by_xpath('//*[@id="schCon"]/div[3]/a/img').click() #확인 클릭
    time.sleep(v_time)
    driver.find_element_by_xpath('//*[@id="searchResult"]/tr/td[1]/a').click() #학현초등학교 클릭
    time.sleep(v_time)
    driver.find_element_by_xpath('//*[@id="exsclyy"]').send_keys('%s' %(v_year - int(student[i]['birthY']) -6)) #학년 선택
    time.sleep(v_time)
    driver.find_element_by_xpath('//*[@id="memberForm"]/div/a[1]').click() #다음 클릭
    time.sleep(v_time)
    print('%s 홈페이지 가입 완료!' % (student[i]['name']))
