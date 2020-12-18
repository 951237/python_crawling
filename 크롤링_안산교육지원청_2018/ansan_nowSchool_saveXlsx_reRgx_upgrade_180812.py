#!/usr/bin/env python3
# -*- coding: utf-8 -*-
#안산 관내 학교 현황 크롤링 #180812

import urllib.request
import bs4
import openpyxl, re
from datetime import date  # 날짜 호출하기

today = date.today().strftime('%y%m%d')

PATH_SAVE = "C:\\Users\\User\\Documents\\coding_python\\crawling\\크롤링_안산교육지원청_2018\\result\\"

url ="http://www.goeas.kr/USR/ORG/MNU9/SchoolList.do?orgType=Z"
html = urllib.request.urlopen(url)
bs_obj = bs4.BeautifulSoup(html, 'html.parser')

all_nowSc = bs_obj.find('table',{'class':'wtl pre-table'})
all_tr = all_nowSc.findAll('tr')

# print(all_tr)

list_tr = []

# 태그를 텍스트로 바꾸어서 리스트로 변환
for tr in all_tr:
    tr_text = tr.text.replace('\n','    ').strip()
    list_tr.append(tr_text)

del list_tr[0]
del list_tr[0]

table = list_tr

print('엑셀파일을 생성합니다. . . ')
wb = openpyxl.Workbook()
sheet = wb['Sheet']

print('데이터를 기록합니다.')
title = ['구분','계', '국립','공립','사립','학생수','교원수']
for _t in range(int(len(title))):
    sheet.cell(row = 1, column = _t+1).value = title[_t]

tableAll = []
divtablereGex = []
for i in range(int(len(table))):
    # \s는 공백 찾기 \s+는 공백이 1개 이상도 찾기
    divreGex = re.compile(r'''(
                            \s+ 
                            )''',re.VERBOSE) #정규식에서 주석 쓰기
    divtable = divreGex.sub(' ',table[i]) #정규식에서 설정한 공백을 ' ' 으로 바꾸기
    divtablereGex.append(divtable) #리스트에 누가하기

for k in range(int(len(divtablereGex))): #리스트이 개수만큼 반복하기
    #글자가 1개 이상, 괄호속의 글자 (1), 가운데 콤마가 있는 단어(5,453)을 찾기
    tablereGex = re.compile(r'''
                            (\w+|\w+\(\w\)|\w*\,\w+)?
                            \s
                            ''',re.VERBOSE) #정규식에서 주석 쓰기
    tablereGexAll = tablereGex.findall(divtablereGex[k])
    tableAll.append(tablereGexAll) #리스트에 누가하기

for rowNum in range(int(len(tableAll))):
    for colNum in range(6):
        # 리스트를 table[0][1]로 표현하기, 중첩된 리스트를 호출할 때 사용
        sheet.cell(row = rowNum + 2, column = colNum + 1).value = tableAll[rowNum][colNum]
wb.save(f'{PATH_SAVE}안산 학교 및 학생 현황_{today}.xlsx')


