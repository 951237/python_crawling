#!/usr/bin/env python3
# -*- coding: utf-8 -*-
#안산관내 초, 중, 고등학교 정보 엑셀 파일로 저장하기

import urllib.request
import bs4
import openpyxl, pprint, os
from datetime import date


def getChoice():
    school = {'1': '초등학교', '2': '중학교', '3': '고등학교', '4': '특수학교'}
    for i in school:
        print('{}{} {}'.format(str(i), '.', school[str(i)]))

    # print(
    #     ' 1 - 초등학교','\n',
    #     '2 - 중학교','\n',
    #     '3 - 고등학교','\n',
    #     '4 - 특수학교','\n'
    # )

    #입력값 받기
    #조건문 : 1 - B, 2 - C, 3 - D, 4 - E
    choice = ''
    s = input('학교를 선택하세요.')
    while True:
        if s == '1':
            choice = 'B'
            break
        elif s == '2':
            choice = 'C'
            break
        elif s == '3':
            choice = 'D'
            break
        elif s == '4':
            choice = 'E'
            break
        elif s == '':
            exit()
        else :
            s = input('학교를 선택하세요.')
    return (choice)
def crawlingWeb(choice):
    print('홈페이지에서 데이터를 다운로드 하고 있습니다...')
    url ='http://www.goeas.kr/USR/ORG/MNU9/SchoolList.do?orgType=%s'%(choice)
    html = urllib.request.urlopen(url)
    bs_obj = bs4.BeautifulSoup(html, 'html.parser')

    all_eleSc = bs_obj.find('table',{'class':'dtl'})
    all_td = all_eleSc.findAll('td')

    list_td = []
    # 태그를 텍스트로 바꾸어서 리스트로 변환
    for td in all_td:
        td_text = td.text.replace('\n','').strip()
        list_td.append(td_text)
    return (list_td)
#1열에 제목을 쓰기
def writeTitle():
    print('데이터를 기록합니다.')
    title = ['설립','학교명','주소','교무실','행정실','팩스']
    for i in range(len(title)):
        sheet.cell(row=1, column = i+1).value = title[i]


def writeContent():
    i = 0
    # 2행부터 데이터 쓰기
    for rowNum in range(int(len(list_td)/6)):
        for colNum in range(6):
            # if len(list_td) < 1 : break
            sheet.cell(row = rowNum+2, column = colNum+1).value = list_td[i]
            i = i + 1
    print('파일 저장중...')
    wb.save('contactSchool.xlsx') #엑셀 파일 저장하기

choice = getChoice()
list_td = crawlingWeb(choice)

print('엑셀파일을 생성중...')
wb = openpyxl.Workbook()
# a = wb.sheetnames
sheet = wb['Sheet']

writeTitle()
writeContent()
