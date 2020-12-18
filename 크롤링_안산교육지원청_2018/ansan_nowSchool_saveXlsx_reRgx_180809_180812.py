#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import urllib.request
import bs4
import openpyxl, re

url ="http://www.goeas.kr/USR/ORG/MNU9/SchoolList.do?orgType=Z"
html = urllib.request.urlopen(url)
bs_obj = bs4.BeautifulSoup(html, 'html.parser')

all_nowSc = bs_obj.find('table',{'class':'wtl pre-table'})
all_tr = all_nowSc.findAll('tr')

# print(all_tr)

list_tr = []


# 태그를 텍스트로 바꾸어서 리스트로 변환
for tr in all_tr:
    tr_text = tr.text.replace('\n','    ').strip()
    list_tr.append(tr_text)

del list_tr[0]
del list_tr[0]

table = list_tr

print('엑셀파일을 생성합니다. . . ')
wb = openpyxl.Workbook()
sheet = wb['Sheet']

print('데이터를 기록합니다.')
title = ['구분','계', '국립','공립','사립','학생수','교원수']
for _t in range(int(len(title))):
    sheet.cell(row = 1, column = _t+1).value = title[_t]

tableAll = []
divtablereGex = []
for i in range(int(len(table))):
    divreGex = re.compile(r'''(
                            \s+
                            )''',re.VERBOSE)
    divtable = divreGex.sub(' ',table[i])
    divtablereGex.append(divtable)

for k in range(int(len(divtablereGex))):
    tablereGex = re.compile(r'''
                            (\w+|\w+\(\w\)|\w*\,\w+)?
                            \s
                            ''',re.VERBOSE)
    tablereGexAll = tablereGex.findall(divtablereGex[k])
    tableAll.append(tablereGexAll)

tableLast = []
for table in tableAll:
    for x in range(int(len(tableAll))):
        tableLast.append(table[x])

_i = 0
for rowNum in range(int(len(tableLast)/6)):
    for colNum in range(6):
        sheet.cell(row = rowNum + 2, column = colNum + 1).value = tableLast[_i]
        _i = _i + 1
wb.save('안산학교 및 학생현황.xlsx')


