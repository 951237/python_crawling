#!/usr/bin/env python3
# -*- coding: utf-8 -*-
#안산교육지원청 과별 업무분장 크롤링 #180731

from urllib import request, parse #url을 파라미터로 받을 때 사용하기 parse
import bs4
from datetime import date #날짜 호출하기
import pandas as pd
import xlwt
import openpyxl



mayday = date.today() #오늘 날짜 저장하기

# ansan_part = {'A':'초등교육지원과','B':'중등교육지원과','C':'평생교육지원과','D':'경영지원과','E':'학교현장지원과','F':'교유시설과'}
name = '초등교육지원과'
print('정보를 불러오는 중....')
url ="http://www.goeas.kr/USR/ORG/MNU6/OrgDetail.do?orgType=A"
html = request.urlopen(url)
bs_obj = bs4.BeautifulSoup(html, 'html.parser')
print('정보 파싱 완료...')

namePart = bs_obj.select('#main_container > div > h4')

print('부서명 수집 완료')
allOffices = bs_obj.select('#main_container > div > div')     #느낀점(뷰티플숍이용하여 간단히 파싱-select 이용)

partHeader = bs_obj.select('#main_container > div > div > table > thead > tr > th')
print('조직도 헤더 수집 완료')
list_partHeader = []
for i in partHeader:
    _i = i.text.lstrip()            #느낀점(왼쪽 여백 삭제)
    list_partHeader.append(_i)

partContent = bs_obj.select('#main_container > div > div > table > tbody > tr > td')
print('조직도 업무내용 수집 완료')
list_partContent = []
for i in partContent:
    _i = i.text.lstrip()
    _i = _i.replace('\r','').replace('\t','').replace('\n','')
    list_partContent.append(_i)

# print(list_partContent)
wb = openpyxl.Workbook()
sheet = wb['Sheet']
print('엑셀 파일 생성 완료')

i_0 = 0
for colNum in range(5):
    sheet.cell(row = 1, column = colNum + 1).value = list_partHeader[i_0]
    i_0 += 1
print('업무분장표 헤더 완성')

i_1 = 0
for rowNum in range(int(len(list_partContent)/5)):  #55개의 데이터를 10줄로 나눠서 뿌리기 #느낀점(데이터를 엑셀 셀에 뿌리기)
    for colNum in range(5):                         # 5열의 데이터를 씀.
        sheet.cell(row = rowNum + 2, column = colNum + 1).value = list_partContent[i_1]
        i_1 += 1                                      # 55개의 데이터를 쓰기 위한 카운터
print('업무분장표 완성')
print('파일 저장중...')
wb.save('pandaText.xlsx')


# df = pd.DataFrame(list_partContent) #내용이 1행으로 나타남.
# print(df)
# writer = pd.ExcelWriter('test_%s.xlsx' %(date.today()))
# df.to_excel(writer, sheet_name='data')
# writer.save()

# output_workbook = xlwt.Workbook
# output_worksheet = output_workbook.add_sheet('data')



# print('작업완료....')