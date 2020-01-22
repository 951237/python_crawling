import pandas as pd
import openpyxl

from bs4 import BeautifulSoup
from selenium import webdriver
import time
import requests

driver = webdriver.Chrome()

# 학교 연락처가 나타난 경기도교육청 페이지
url = 'http://www.goe.go.kr/edu/school/selectSchoolList.do?menuId=270151220143616&programId=PGM_1000000004&top_m=1&left_m=5&lm_1=6'

print('사이트에 접속중..')
driver.get(url)
time.sleep(0.25)

### 함수 - 학교정보 가져오기(Td)


def getSchoolInfoTd(p_list):
    k = 1
    for j in range(3, 13):
        tds = driver.find_element_by_xpath('//*[@id="contents"]/div[3]/div[2]/table/tbody').find_elements_by_tag_name(
            'td')  # 표 안에서 행 찾기
        page = driver.find_element_by_xpath('//*[@id="contents"]/div[3]/div[3]/a[%s]' % (j))
        numPage = page.get_attribute('title')
        for td in tds:
            i = td.text
            p_list.append(i)  # 리스트에 저장하기

        print('%s 처리중....' % (numPage))
        #         print('%s 페이지 처리중....' %(p_k*10 + k))
        page.click()
        time.sleep(0.5)
        k += 1
    return p_list


### 함수 - 리스트의 내용을 파일로 저장하기

def saveTextFile(p_list):
    out = open('schoolContactAll_181109-1.txt', 'w')

    for i in p_list:
        print(i, file=out)

    out.close()


### 함수 - 마지막 표의 학교 정보를 리스트로 저장하기(Td)

def getSchoolInfoOneTd(p_list):
    tds = driver.find_element_by_xpath('//*[@id="contents"]/div[3]/div[2]/table/tbody').find_elements_by_tag_name(
        'td')  # 표 안에서 행 찾기

    for td in tds:
        i = td.text
        p_list.append(i)
    return p_list


### 리스트 엑셀파일로 저장하기

#### 함수 - 엑셀파일 행제목 쓰기

def writeTitle():
    print('데이터를 기록합니다.')
    title = ['지역', '학급', '설립', '학교명', '주소', '전화번호']
    for i in range(len(title)):
        sheet.cell(row=1, column=i + 1).value = title[i]


#### 함수 - 엑셀파일 내용 쓰기

def writeContent(p_list):
    i = 0
    # 2행부터 데이터 쓰기
    for rowNum in range(int(len(p_list) / 6)):
        for colNum in range(6):
            # if len(list_td) < 1 : break
            sheet.cell(row=rowNum + 2, column=colNum + 1).value = p_list[i]
            i = i + 1
    print('파일 저장중...')
    wb.save('kyunggiSchoolCantacAllTd_181110.xlsx')  # 엑셀 파일 저장하기


### 실행 - 표의 학교 정보를 리스트로 저장하기

k = 0
listSchoolContac = []
while k < 48:
    trs = driver.find_element_by_xpath('//*[@id="contents"]/div[3]/div[2]/table/tbody').find_elements_by_tag_name(
        'tr')  # 표 안에서 행 찾기
    if len(trs) > 0:
        getSchoolInfoTd(listSchoolContac)
        k += 1
        time.sleep(0.5)

    else:
        break

# 파일저장 함수 실행
print('마지막 페이지 스크랩...')
getSchoolInfoOneTd(listSchoolContac)

print('엑셀파일 저장중...')
wb = openpyxl.Workbook()
sheet = wb['Sheet']

writeTitle()  # 엑셀파일 타이틀 쓰기
writeContent(listSchoolContac)  # 엑셀파일 내용 쓰기

print('작업완료')