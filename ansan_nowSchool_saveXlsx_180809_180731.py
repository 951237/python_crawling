#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import urllib.request
import bs4
import openpyxl

url ="http://www.goeas.kr/USR/ORG/MNU9/SchoolList.do?orgType=Z"
html = urllib.request.urlopen(url)
bs_obj = bs4.BeautifulSoup(html, 'html.parser')

all_nowSc = bs_obj.find('table',{'class':'wtl pre-table'})
all_tr = all_nowSc.findAll('tr')

# print(all_tr)

list_tr = []


# 태그를 텍스트로 바꾸어서 리스트로 변환
for tr in all_tr:
    tr_text = tr.text.replace('\n','    ').strip()
    list_tr.append(tr_text)

del list_tr[0]
del list_tr[0]

table = list_tr

print('엑셀파일을 생성합니다. . . ')
wb = openpyxl.Workbook()
sheet = wb['Sheet']

print('데이터를 기록합니다.')
title = ['구분','계', '국립','공립','사립','학생수','교원수']
for _t in range(int(len(title))):
    sheet.cell(row = 1, column = _t+1).value = title[_t]

for i in range(int(len(table))):
    sheet.cell(row = i + 2, column = 1).value = table[i]

wb.save('ansan_school&stu.xlsx')

